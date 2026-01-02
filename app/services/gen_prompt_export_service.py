import os
import json
import glob
from datetime import datetime
from openpyxl import load_workbook
from app.config import settings

EXPORT_XLSX_DIR = settings.BASE_DIR / "exports_xlsx"
PROMPTS_JSON_DIR = settings.BASE_DIR / "exports_prompts_json"

os.makedirs(PROMPTS_JSON_DIR, exist_ok=True)

def save_prompts_data(session_id: str, data: dict):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    json_path = PROMPTS_JSON_DIR / f"prompts_{session_id}_{timestamp}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4)
    
    search_pattern = EXPORT_XLSX_DIR / f"session_{session_id}_*.xlsx"
    files = glob.glob(str(search_pattern))
    
    if not files:
        print(f"Warning: No Excel file found for session {session_id}. Prompts saved to JSON only.")
        return json_path, None
        
    latest_file = max(files, key=os.path.getctime)
    
    try:
        wb = load_workbook(latest_file)
    except Exception as e:
        print(f"Error loading Excel: {e}")
        return json_path, None

    if "Prompts" in wb.sheetnames:
        del wb["Prompts"]
    
    ws = wb.create_sheet("Prompts")
    headers = ["Screen", "developer_style", "design_style", "copy-style"]
    ws.append(headers)

    for screen in data.get("screens", []):
        prompts = screen.get("prompts", {})
        
        ws.append([
            screen.get("screen_name"),
            prompts.get("developer"),
            prompts.get("designer"),
            prompts.get("copywriter")
        ])

    wb.save(latest_file)
    
    return json_path, latest_file
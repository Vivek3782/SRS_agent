from openpyxl import load_workbook
import os
import json
import glob
from datetime import datetime
from openpyxl import Workbook
from app.config import settings
from pathlib import Path

EXPORT_XLSX_DIR = settings.BASE_DIR / "exports_xlsx"
EXPORT_JSON_DIR = settings.BASE_DIR / "exports_json"
ESTIMATED_PAGES_DIR = settings.BASE_DIR / "estimated_pages_json"
EXPORT_IMAGES_DIR = settings.BASE_DIR / "exports_branding_images"

os.makedirs(EXPORT_XLSX_DIR, exist_ok=True)
os.makedirs(EXPORT_JSON_DIR, exist_ok=True)
os.makedirs(ESTIMATED_PAGES_DIR, exist_ok=True)
os.makedirs(EXPORT_IMAGES_DIR, exist_ok=True)


def get_session_xlsx_path(session_id: str) -> Path:
    """
    Finds the latest Excel file for a session or generates a new path.
    """
    search_pattern = EXPORT_XLSX_DIR / f"session_{session_id}_*.xlsx"
    files = glob.glob(str(search_pattern))
    if files:
        return Path(max(files, key=os.path.getctime))

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return EXPORT_XLSX_DIR / f"session_{session_id}_{timestamp}.xlsx"


def save_to_excel(session_id: str, history: list):
    """
    Saves/Appends conversation history to the 'exports_xlsx' folder.
    """
    filepath = get_session_xlsx_path(session_id)

    if filepath.exists():
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
        # Remove default sheet if we are creating new
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # --- Manage "Conversation Log" Sheet ---
    sheet_name = "Conversation Log"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(["Question", "Answer", "Timestamp", "Customer ID"])

    for item in history:
        q = item.get("question") if isinstance(item, dict) else item.question
        a = item.get("answer") if isinstance(item, dict) else item.answer
        t = item.get("timestamp") if isinstance(item, dict) else item.timestamp
        s = item.get("session_id") if isinstance(
            item, dict) else item.session_id
        ws.append([q, a, t, s])

    wb.save(filepath)
    return filepath


def save_requirements(session_id: str, requirements: dict):
    """
    Saves the final requirements JSON to the 'exports_json' folder.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"requirements_{session_id}_{timestamp}.json"

    # Save to JSON Directory
    filepath = EXPORT_JSON_DIR / filename

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(requirements, f, indent=4, ensure_ascii=False)

    return filepath


def get_latest_requirements_file(session_id: str):
    """
    Reads the SOURCE requirements from 'exports_json'.
    """
    search_pattern = EXPORT_JSON_DIR / f"requirements_{session_id}_*.json"
    files = glob.glob(str(search_pattern))

    if not files:
        return None, None

    latest_file = max(files, key=os.path.getctime)

    try:
        with open(latest_file, "r", encoding="utf-8") as f:
            data = json.load(f)
            return latest_file, data
    except Exception as e:
        print(f"Error reading file {latest_file}: {e}")
        return latest_file, None


def save_estimated_sitemap(session_id: str, sitemap_data: dict):
    """
    Saves the generated sitemap to the NEW 'estimated_pages_json' folder.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"sitemap_{session_id}_{timestamp}.json"

    filepath = ESTIMATED_PAGES_DIR / filename

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(sitemap_data, f, indent=4, ensure_ascii=False)

    return filepath


def append_screens_to_excel(session_id: str, sitemap_data: dict):
    # 1. Find the existing Excel file (Scanning EXPORT_XLSX_DIR)
    search_pattern = EXPORT_XLSX_DIR / f"session_{session_id}_*.xlsx"
    files = glob.glob(str(search_pattern))

    if not files:
        print("No Excel file found to append screens.")
        return None

    latest_file = max(files, key=os.path.getctime)

    # 2. Load Workbook
    wb = load_workbook(latest_file)

    # 3. Manage "Screens" Sheet
    if "Screens" in wb.sheetnames:
        del wb["Screens"]

    ws = wb.create_sheet("Screens")

    # 4. Write Headers
    headers = ["Screen Name", "Complexity", "Notes", "Description", "Features"]
    ws.append(headers)

    # 5. Write Data
    for page in sitemap_data.get("pages", []):
        # Handle list of features for CSV-like cell
        features_str = ", ".join(page.get("features", []))

        ws.append([
            page.get("name"),
            page.get("complexity", "Medium"),
            page.get("notes", ""),
            page.get("description", ""),
            features_str
        ])

    # 6. Save
    wb.save(latest_file)
    return latest_file


def delete_estimated_sitemap(session_id: str):
    """
    Deletes the estimated sitemap from the NEW 'estimated_pages_json' folder.
    """
    search_pattern = ESTIMATED_PAGES_DIR / f"sitemap_{session_id}_*.json"
    files = glob.glob(str(search_pattern))

    if not files:
        return None

    latest_file = max(files, key=os.path.getctime)

    try:
        with open(latest_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        os.remove(latest_file)
        return data
    except Exception as e:
        print(f"Error deleting file {latest_file}: {e}")
        return None


# for branding
BRANDING_JSON_DIR = settings.BASE_DIR / "exports_branding_json"
BRANDING_XLSX_DIR = settings.BASE_DIR / "exports_branding_xlsx"

os.makedirs(BRANDING_JSON_DIR, exist_ok=True)
os.makedirs(BRANDING_XLSX_DIR, exist_ok=True)


def save_branding_files(session_id: str, state_data: dict, only_json: bool = False):
    """
    Saves both JSON profile and updates the session Excel with branding info.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 1. Save JSON Profile
    json_filename = f"branding_profile_{session_id}_{timestamp}.json"
    json_path = BRANDING_JSON_DIR / json_filename

    profile_data = state_data.get("profile", {})
    if isinstance(profile_data, object) and hasattr(profile_data, "model_dump"):
        profile_data = profile_data.model_dump()

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(profile_data, f, indent=4)

    if only_json:
        return json_path, None

    # 2. Update Excel Summary & Transcript
    filepath = get_session_xlsx_path(session_id)

    if filepath.exists():
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # --- Sheet 1: Profile Summary (The "SRS" format) ---
    sheet_summary = "Company Profile"
    if sheet_summary in wb.sheetnames:
        del wb[sheet_summary]

    ws_summary = wb.create_sheet(sheet_summary, 0)  # Place at first
    ws_summary.append(["Requirement Category", "Collected Information"])

    for field, value in profile_data.items():
        if value:
            display_name = field.replace("_", " ").title()
            if isinstance(value, list):
                display_value = ", ".join(map(str, value))
            elif isinstance(value, dict):
                display_value = ", ".join(
                    [f"{k}: {v}" for k, v in value.items()])
            else:
                display_value = str(value)
            ws_summary.append([display_name, display_value])

    # --- Sheet 2: Branding Chat Transcript ---
    sheet_chat = "Branding Chat"
    if sheet_chat in wb.sheetnames:
        ws_chat = wb[sheet_chat]
    else:
        ws_chat = wb.create_sheet(sheet_chat)
        ws_chat.append(["Question", "User Answer"])

    # We only append NEW history items?
    # Actually, for simplicity, let's just refresh history or store last saved index.
    # But since history in state is cumulative, clearing and re-writing is safer.
    if sheet_chat in wb.sheetnames:
        del wb[sheet_chat]
        ws_chat = wb.create_sheet(sheet_chat)
        ws_chat.append(["Question", "User Answer"])

    history = state_data.get("history", [])
    for turn in history:
        q = turn.get("question") if isinstance(turn, dict) else turn.question
        a = turn.get("answer") if isinstance(turn, dict) else turn.answer
        ws_chat.append([q, a])

    wb.save(filepath)
    return json_path, filepath


BRANDING_JSON_DIR = settings.BASE_DIR / "exports_branding_json"


def get_branding_export(session_id: str) -> dict | None:
    """
    Checks if a completed Branding Profile exists for this session.
    Returns the profile data dict if found, otherwise None.
    """

    search_pattern = BRANDING_JSON_DIR / \
        f"branding_profile_{session_id}_*.json"
    files = glob.glob(str(search_pattern))

    if not files:
        return None

    # Get the latest one if multiple exist
    latest_file = max(files, key=os.path.getctime)

    try:
        with open(latest_file, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None
